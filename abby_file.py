#!C:\Python27
# -*- coding: utf-8 -*-
import sys
import re
from openpyxl import load_workbook, Workbook
from utils import get_unicode, find_nth, convert_to_float


# ROW PARSERS
class BaseParser():
    """Provides row and context members to derived parsers, and a standard
    method accepts to check if a row can be parsed for a specific derived
    parser.

    Derived parsers will parse row and modify context with the results."""

    def __init__(self, row=None, context=None):
        self.row = row
        self.context = context

    def accepts(self):
        """Check if a row can be parsed by the derived parser.

        Uses up to 3 conditions stored on derived parsers in load_conditions
        method to check if the row is recognized:
            1. Containing a substring
            2. Length of row (number of list items in row)
            3. Pattern matching
        """

        # load derived parsers conditions to check
        self.load_conditions()

        # substring contained condition
        if self.row_substring:
            substring_cond = self.row_substring in self.row[0]
        else:
            substring_cond = True

        # length of row condition
        if self.row_length:
            len_cond = len(self.row) == self.row_length
        else:
            len_cond = True

        # pattern matching condition
        if self.row_pattern:
            pattern_cond = self._re_match(self.row_pattern, self.row[0])
        else:
            pattern_cond = True

        return substring_cond and len_cond and pattern_cond

    def _re_match(self, pattern, string):
        """Checks if string matches pattern."""
        RV = False

        match_obj = re.match(pattern, string, re.U)

        if match_obj:
            RV = True

        return RV


class TitleParser(BaseParser):
    """Parse title rows getting id_title and desc_title."""

    def load_conditions(self):
        """Load accepting conditions for BaseParser.accepts() method."""
        self.row_substring = u"T\xcdTULO"
        self.row_length = 1
        self.row_pattern = None

    def parse(self):
        """Parse data from row"""

        # modify context with parsing results
        self.context.id_title = self._get_id_title()
        self.context.desc_title = self._get_desc_title()

        # modify context with type of row parsed
        self.context.row_type = "title"

    def _get_id_title(self):
        """Parse id_title from first substring before first dot position."""

        # find first dot
        i_dot = self.row[0].find(".")

        # take title word out and strip whitespaces
        RV = self.row[0][:i_dot].replace(u'T\xcdTULO', "").strip()

        return RV

    def _get_desc_title(self):
        """Parse desc_title from substring starting after first dot position."""

        # find first dot
        i_dot = self.row[0].find(".")

        # keeps substring from first dot to end
        substring = self.row[0][i_dot + 1:]

        # extract substring that matches pattern
        pattern = "[A-Z][A-Z\s]{1,}"
        RV = re.search(pattern, substring, re.U).group().strip()

        return RV


class Subt1Parser(BaseParser):
    """Parse first level of subtitle rows getting id_subt1 and desc_subt1."""

    def load_conditions(self):
        """Load accepting conditions for BaseParser.accepts() method."""
        self.row_substring = None
        self.row_length = 1
        self.row_pattern = "^[a-z][)][A-Z\s]{1,}"

    def parse(self):
        """Parse data from row."""

        # modify context with parsing results
        self.context.id_subt1 = self._get_id_subt1()
        self.context.desc_subt1 = self._get_desc_subt1()

        # modify context with type of row parsed
        self.context.row_type = "subt1"

    def _get_id_subt1(self):
        """Parse id_subt1 from first substring before first ")" position."""

        # find first parenthesis
        i_parenthesis = self.row[0].find(")")

        # take substring up to first parenthesis and strip whitespaces
        RV = self.row[0][:i_parenthesis].strip()

        return RV

    def _get_desc_subt1(self):
        """Parse desc_subt1 from substring starting after first ")" position."""

        # find first parenthesis
        i_parenthesis = self.row[0].find(")")

        # keeps substring from first parenthesis to end
        substring = self.row[0][i_parenthesis + 1:]

        # extract substring that matches pattern
        pattern = "[A-Z][A-Z\s]{1,}"
        RV = re.search(pattern, substring, re.U).group().strip()

        return RV


class Subt2Parser(BaseParser):
    """Parse 2nd level of subtitle rows getting id_subt2 and desc_subt2."""

    def load_conditions(self):
        """Load accepting conditions for BaseParser.accepts() method."""
        self.row_substring = None
        self.row_length = 1
        self.row_pattern = "^[0-9][\.]"

    def parse(self):
        """Parse data from row."""

        # modify context with parsing results
        self.context.id_subt2 = self._get_id_subt2()
        self.context.desc_subt2 = self._get_desc_subt2()

        # modify context with type of row parsed
        self.context.row_type = "subt2"

    def _get_id_subt2(self):
        """Parse id_subt2 from first substring before first dot position."""

        # find first dot
        i_dot = self.row[0].strip().find(".")

        # take stripped substring up to first dot
        RV = self.row[0].strip()[:i_dot]

        return RV

    def _get_desc_subt2(self):
        """Parse desc_subt2 from substring starting after first dot position."""

        # find first dot
        i_dot = self.row[0].find(".")

        # keeps substring from first dot to end
        substring = self.row[0][i_dot + 1:]

        # extract substring that matches pattern
        pattern = "[A-Z].{1,}"
        RV = re.search(pattern, substring, re.U).group().strip()

        return RV


class AgValuesParser(BaseParser):
    """Parse aggregated values rows and set context variables in aggregated
    codes or descriptions according with values level of aggregation."""

    def load_conditions(self):
        """Load accepting conditions for BaseParser.accepts() method."""
        self.row_substring = u"Valor total:"
        self.row_length = 1
        self.row_pattern = None

    def parse(self):
        """Parse data from row and set context variables for aggregated values
        values."""

        # if last row was a title, we are the top level of aggregation
        # all subtitles will be set to aggregated codes or descriptions
        if self.context.row_type == "title":

            # first subtitle
            self.context.id_subt1 = 0
            self.context.desc_subt1 = "Todos"

            # second subtitle
            self.context.id_subt2 = 0
            self.context.desc_subt2 = "Todos"

        # if last row was a "first subtitle", we are at the second level of
        # aggregation, next subtitle level will be set to agg code and desc
        if self.context.row_type == "subt1":

            # second subtitles
            self.context.id_subt2 = 0
            self.context.desc_subt2 = "Todos"

        # in any case, we are at higher agg level than "product"
        self.context.id_product = 0
        self.context.tariff_number = u"NA"
        self.context.desc_product = u"Todos"
        self.context.product_units = u"NA"

        # no country or quantity data with aggregated values
        self.context.id_country = 0
        self.context.desc_country = "Todos"
        self.context.quantity = [None, None]

        # modify context with parsing results
        self.context.year = self._get_year()
        self.context.value = self._get_value()

        # modify context with type of row parsed
        self.context.row_type = "agg_values"

    def _get_year(self):
        """Parse years based on ":", "m$n" and ";" delimiters"""
        RV = []

        # first year, between ":" and "m$n"
        yr_start = self.row[0].find(":") + 1
        yr_end = self.row[0].find("m$n")

        # strip and convert to int
        RV.append(int(self.row[0][yr_start:yr_end].strip()))

        # second year, between ";" and last "m$n"
        yr_start = self.row[0].find(";") + 1
        yr_end = self.row[0].rfind("m$n")

        # strip and convert to int
        RV.append(int(self.row[0][yr_start:yr_end].strip()))

        return RV

    def _get_value(self):
        """Parse values based on "m$n", ";" and ")" delimiters"""
        RV = []

        # first value, between "m$n" and ";"
        # add +4 because "m$n" delimiter has 3 chars, plus 1 white space
        value_start = self.row[0].find("m$n") + 4
        value_end = self.row[0].find(";")

        # get substring, convert to float and append
        str_value = self.row[0][value_start:value_end]
        float_value = convert_to_float(str_value)
        RV.append(float_value)

        # second value, between last "m$n" and ")"
        # add +4 because "m$n" delimiter has 3 chars, plus 1 white space
        value_start = self.row[0].rfind("m$n") + 4
        value_end = self.row[0].find(")")

        # get substring, convert to float and append
        str_value = self.row[0][value_start:value_end]
        float_value = convert_to_float(str_value)
        RV.append(float_value)

        return RV


class Head1Parser(BaseParser):
    """Parse head of table with tariff number row getting id_product,
    tariff_number, desc_product and product_units."""

    def load_conditions(self):
        """Load accepting conditions for BaseParser.accepts() method."""
        self.row_substring = "Tarifa"
        self.row_length = 1
        self.row_pattern = "^[0-9].{1,}Tarifa.{1,}:"

    def parse(self):
        """Parse data from row."""

        # modify context with parsing results
        self.context.id_product = self._get_id_product()
        self.context.tariff_number = self._get_tariff_number()
        self.context.desc_product = self._get_desc_product()
        self.context.product_units = self._get_product_units()

        # modify context with type of row parsed
        self.context.row_type = "tbl_head1"

    def _get_id_product(self):
        """Parse id_product from substring using min index of several possible
        delimiters."""

        index_list = [self.row[0].strip().find("."),
                      self.row[0].strip().find("-"),
                      self.row[0].strip().find("(")]

        # remove -1, as could lead to error finding the minimum index
        try:
            index_list.remove(-1)
        except:
            pass

        # get min index identified
        min_index = min(index_list)

        # if greater than 1, is used
        if min_index > 1:
            index = min_index
        # if not, uses 1 (this avoid using 0)
        else:
            index = 1

        return self.row[0].strip()[:index]

    def _get_tariff_number(self):
        """Parse tariff_number from substring between "Tarifa" and ")"."""

        # if several tariff numbers apply, state it in the same way
        if u"varios y no tarifados" in self.row[0]:
            RV = "varios y no tarifados"

        # take indexes using "Tarifa" and ")"
        else:
            start = self.row[0].find(u"Tarifa") + 6
            end = self.row[0].find(u")")
            RV = self.row[0][start:end].strip()

        return RV

    def _get_desc_product(self):
        """Parse desc_product from substring starting at second dot and ending
        at last comma."""

        ### check that there is a dot at beggining of row
        # find first dot
        i_dot = self.row[0].find(".")

        # find both parenthesis indexes
        i_parenthesis = [self.row[0].find("("), self.row[0].find(")")]

        # remove -1, as could lead to error finding the minimum index
        try:
            i_parenthesis.remove(-1)
        except:
            pass

        # get min index identified
        min_index = min(i_parenthesis)

        # if i_dot is less than min_index of both parenthesis, there is a dot
        # before them, and thus there is a "dot at beggining of row"
        # take second dot in row, avoiding the first one
        if i_dot < min_index:
            i_start = find_nth(self.row[0], ".", 2) - 1

        # if there is not dot at the begining, take first dot in row
        else:
            i_start = find_nth(self.row[0], ".", 1) - 1

        # find end index by last comma
        i_end = self.row[0].rfind(",")

        ### take substring with indexes found before
        substring = self.row[0][i_start:i_end]

        # regex pattern
        pattern = "[A-Z].{1,}"

        # try match the pattern
        try:
            RV = re.search(pattern, substring, re.U).group().strip()
        # if not possible, returns parsing error
        except:
            RV = "Parsing error"

        return RV

    def _get_product_units(self):
        """Parse product_units from substring starting at last comma and
        ending at last ":"."""

        # find last comma
        i_start = self.row[0].rfind(",") + 1

        # find last ":"
        i_end = self.row[0].rfind(":")

        # keeps substring between last comma and ":"
        RV = self.row[0][i_start:i_end].strip()

        # normalize product_units denomination
        RV = self._normalize_product_units(RV)

        return RV

    def _normalize_product_units(self, product_units):
        """Normalize string for "kilogramos" unit, if that is the product_unit
        passed."""
        RV = product_units

        # uses the long form of the unit
        if get_unicode(product_units.strip()) == u"Kg.":
            RV = u"kilogramos"

        return RV


class Head2Parser(BaseParser):
    """TODO: Still need to be implemented. Class for parsing exports head of
    table that do not have tariff numbers."""

    def load_conditions(self):
        """Load accepting conditions for BaseParser.accepts() method."""
        self.row_substring = None
        self.row_length = 1
        self.row_pattern = "^[0-9].{1,}:"


class TblRowParser(BaseParser):
    """Parse tbl_row data (the actual data, at least aggregation level) from
    a 5 cells row.

    Cells contain:
        1. Country name (desc_country)
        2. 1945 quantity
        3. 1946 quantity
        4. 1945 value
        5. 1946 value"""

    def load_conditions(self):
        """Load accepting conditions for BaseParser.accepts() method."""
        self.row_substring = None
        self.row_length = 5
        self.row_pattern = None

    def parse(self):
        """Parse data from row."""

        # modify context with parsing results
        self.context.id_country = None  # TODO: there will be ids for countries
        self.context.desc_country = self._get_desc_country()
        self.context.year = self._get_year()
        self.context.quantity = self._get_quantity()
        self.context.value = self._get_value()

        # modify context with type of row parsed
        self.context.row_type = "tbl_row"

    def _get_desc_country(self):
        """Parse desc_country from first cell of row"""

        # checks if is missing
        if not self.row[0]:
            RV = "Missing error"

        # checks if desc_country refers to all countries
        elif u"Total" in self.row[0] or u"total" in self.row[0]:
            RV = u"Todos"

        # if there is an individual country name, get rid of dots and take it
        else:
            RV = self.row[0].replace(".", "").strip()

        return RV

    def _get_year(self):
        """For this document, years are always 1945 and 1946. This will change
        for any other document."""

        RV = [u"1945", u"1946"]

        return RV

    def _get_quantity(self):
        """Parse quantities from tbl_row, first two cells after country name."""
        quantity_1945 = None
        quantity_1946 = None

        # if cant convert to float, is missing
        try:
            quantity_1945 = convert_to_float(self.row[1])
        except:
            pass

        # if cant convert to float, is missing
        try:
            quantity_1946 = convert_to_float(self.row[2])
        except:
            pass

        return [quantity_1945, quantity_1946]

    def _get_value(self):
        """Parse values from tbl_row, last two cells in row."""
        value_1945 = None
        value_1946 = None

        # if cant convert to float, is missing
        try:
            value_1945 = convert_to_float(self.row[3])
        except:
            pass

        # if cant convert to float, is missing
        try:
            value_1946 = convert_to_float(self.row[4])
        except:
            pass

        return [value_1945, value_1946]


class IgnoreRow(BaseParser):
    """Take rows that must be ignored."""

    def accepts(self):
        """Override base method. Checks for special conditions to ignore row."""

        # if row is None, is ignored
        if (not self.row) or (not self.row[0]):
            return True

        # if row is page continuation of splitted table, ignores it
        # information contained in this row, is already in context
        continuation_row = u"(Conclusi\xf3n)" in self.row[0]

        # if row is empty, is ignored
        empty_row = u"" == self.row[0]

        return continuation_row or empty_row

    def parse(self):
        """Just declare type row as "ignore"."""

        self.context.row_type = "ignore"


class NoneImportParser(BaseParser):
    """Generate empty records for tables with no records. These show a string
    stating there's no imports."""

    def load_conditions(self):
        """Load accepting conditions for BaseParser.accepts() method."""
        self.row_substring = u"Sin importaci\xf3n"
        self.row_length = 1
        self.row_pattern = None

    def parse(self):
        """Set context for empty records."""

        self.context.id_country = 0
        self.context.desc_country = u"Todos"
        self.context.year = [1945, 1946]
        self.context.quantity = [u"NA", u"NA"]
        self.context.value = [u"NA", u"NA"]

        # modify context with type of row
        self.context.row_type = "noneImport"


class Head1IniPart(BaseParser):
    """Hold first part of head1 (when is splitted into two rows) in
    context.last_row variable to merge later with the second part."""

    def load_conditions(self):
        """Load accepting conditions for BaseParser.accepts() method."""
        self.row_substring = None
        self.row_length = 1
        self.row_pattern = "^[0-9].{1,}Tarifa.{1,}"

    def parse(self):
        """Take initial part of head1 row."""

        # modify context storing partial row
        self.context.last_row = self.row[0]

        # modify context with type of row holded
        self.context.row_type = "tbl_head1_ini_part"


class Head1FinalPart(BaseParser):
    """Merge second part of head1 row with first one (stored in context) and
    use Head1Parser to parse the merged row."""

    def load_conditions(self):
        """Load accepting conditions for BaseParser.accepts() method."""
        self.row_substring = None
        self.row_length = 1
        self.row_pattern = ".{1,}:"

    def parse(self):
        """Merge both parts of head1 row and call Head1Parser."""

        # forma la row completa de tipo Head1
        tbl_head1_row = [get_unicode(self.context.last_row + u" " +
                         self.row[0])]
        tbl_head1_row = [tbl_head1_row[0].strip()]

        # check that row is accepted and use parser
        if Head1Parser().accepts(tbl_head1_row):
            Head1Parser(tbl_head1_row, self.context).parse()
        # if not accepted, print an error
        else:
            print tbl_head1_row
            print "Ocurrio un error con un Head1Parser partido!"

        # declara cual fue el ultimo row type procesado
        self.context.row_type = "tbl_head1_final_part"


# INTERNAL CLASSES
class StatsBook1Context():
    """Provides context for Stats Book 1. Variables in context are mainly
    fields of parsed database. Last ones are for internal use of parsers."""

    def __init__(self):

       # title
        self.id_title = None
        self.desc_title = None

        # first level subtitle
        self.id_subt1 = None
        self.desc_subt1 = None

        # second level subtitle
        self.id_subt2 = None
        self.desc_subt2 = None

        # head product table
        self.id_product = None
        self.tariff_number = None
        self.desc_product = None
        self.product_units = None

        # data inside product table
        self.id_country = None
        self.desc_country = None
        self.year = []
        self.quantity = []
        self.value = []

        # context control variables for parsers
        self.row_type = ""
        self.type_last_row = ""
        self.last_row = ""


class StatsBook1RecordsBuilder():
    """Build records from a StatsBook1Context instance."""

    def __init__(self, context):
        self.context = context

    def build_records(self):
        """Build records using the context and modifies it."""

        new_records = []

        # only if row_type is one with data, records can be built
        if self.context.row_type == "agg_values" or \
                self.context.row_type == "tbl_row":

            # each value has a year and quantity, the other variables are equal
            i = 0
            for value in self.context.value:

                new_record = dict()

                # title
                new_record["id_title"] = self.context.id_title
                new_record["desc_title"] = self.context.desc_title

                # first level subtitle
                new_record["id_subt1"] = self.context.id_subt1
                new_record["desc_subt1"] = self.context.desc_subt1

                # second level subtitle
                new_record["id_subt2"] = self.context.id_subt2
                new_record["desc_subt2"] = self.context.desc_subt2

                # head product table
                new_record["id_product"] = self.context.id_product
                new_record["tariff_number"] = self.context.tariff_number
                new_record["desc_product"] = self.context.desc_product
                new_record["product_units"] = self.context.product_units

                # data inside product table
                new_record["id_country"] = self.context.id_country
                new_record["desc_country"] = self.context.desc_country
                new_record["year"] = self.context.year[i]
                new_record["quantity"] = self.context.quantity[i]
                new_record["value"] = value

                # adds new record
                new_records.append(new_record)
                i += 1

        # update last row_type
        self.context.row_type = str(self.context.row_type)

        return new_records


class AbbyParser():
    """Parse rows of ABBY output and build records from them.

    AbbyParser object needs a list of parsers and a context to be built.
    Iterates through list of parsers looking for one that accepts the row, then
    parser modifies the context based on extracted information of row and
    AbbyParser build records from the context and return them."""

    def __init__(self, parsers, context, records_builder):
        self.context = context()
        self.parsers = parsers
        self.records_builder = records_builder

    def parse_row(self, row):
        """Main method. Parse a row modifying context an build records."""

        # iterate through parsers
        for parser_class in self.parsers:
            if parser_class(row).accepts():

                # parse row and modify context with results
                parser = parser_class(row, self.context)
                parser.parse()
                break

        # TODO: What happens if no parser accepts the row???

        # yield any new records that can be built, after row was parsed
        return self.records_builder(self.context).build_records()


# USER CLASSES
class AbbyFile():
    """Takes a workbook with a single sheet that is ABBY ocr output from an old
    stats book. Parse all rows of the worksheet building database records.

    Uses AbbyParser class to handle parsing wich needs parsers, context and
    record builder to do it."""

    # DATA
    PARSERS = [IgnoreRow, NoneImportParser, Head1Parser, Head1IniPart,
               AgValuesParser, Head1FinalPart, TblRowParser, Head2Parser,
               TitleParser, Subt1Parser, Subt2Parser]

    CONTEXTS = [StatsBook1Context]

    RECORDS_BUILDERS = [StatsBook1RecordsBuilder]

    FIELDS = ["id_title",
              "desc_title",
              "id_subt1",
              "desc_subt1",
              "id_subt2",
              "desc_subt2",
              "id_product",
              "tariff_number",
              "desc_product",
              "product_units",
              "id_country",
              "desc_country",
              "year",
              "quantity",
              "value"]

    def __init__(self, wb):
        self.wb = wb

    # PUBLIC
    def get_records(self):
        """Read all rows of workbook and parse them looking to build database
        records from them. Yield records as soon as they are built."""

        # load active sheet of wb
        ws = self.wb.get_active_sheet()

        # create AbbyParser instance
        ap = AbbyParser(self.PARSERS, self.CONTEXTS[0], self.RECORDS_BUILDERS[0])

        # iterate through all abby_file wb rows
        for row in ws.iter_rows():
            cells_values = []

            # iterate through all cells in row
            for cell in row:
                # add cell value
                cells_values.append(get_unicode(cell.internal_value))

            # checks if list of cell values not empty
            if not self._empty(cells_values):
                # remove any empty cell that might be at the end of row
                cells_values = self._remove_lasts_none(cells_values)

                # parse all posible records from row (list of cell values)
                record_lines = ap.parse_row(cells_values)

                # yields any record built from row
                for record in record_lines:
                    yield record

    # PRIVATE
    def _empty(self, values_list):
        """True if values_list is empty."""
        RV = True

        for value in values_list:
            if value is not None:
                RV = False

        return RV

    def _remove_lasts_none(self, values_list):
        """Delete None values at the end of a list."""
        RV = values_list

        while RV[-1] is None:
            del RV[-1]

        return RV

# DATA
ABBY_FILE_NAME = "abby_file.xlsx"
ABBY_PARSED_FILE_NAME = "abby_parsed.xlsx"


# USER METHODS
def write_ws(ws, record, fields):
    """Add a record to a worksheet."""
    new_row = []

    # extract data with field keys from record
    for field in fields:
        new_row.append(record[field])

    # add new row to worksheet
    ws.append(new_row)


def scrape_abby_file(wb_abby_name=None, wb_abby_parsed_name=None):
    """Takes an abby output excel file and returns a database formatted excel
    file with records built from it."""

    # if not wb names passed, defaults name are used
    wb_abby_name = wb_abby_name or ABBY_FILE_NAME
    wb_abby_parsed_name = wb_abby_parsed_name or ABBY_PARSED_FILE_NAME

    # loads abby file
    wb_abby = load_workbook(filename=wb_abby_name, use_iterators=True)
    abby_file = AbbyFile(wb_abby)

    # creates new excel sheet to store new records
    wb_parsed = Workbook(optimized_write=True)
    ws_parsed = wb_parsed.create_sheet()

    # write field names
    ws_parsed.append(abby_file.FIELDS)

    # write every record parsed in the database formatted excel sheet
    for record in abby_file.get_records():
        write_ws(ws_parsed, record, abby_file.FIELDS)

    # save database formatted excel with parsed records
    wb_parsed.save("abby_parsed.xlsx")


# executes main routine
if __name__ == '__main__':

    # if parameters are passed, use them
    input_file = None
    output_file = None

    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]

    # main call
    scrape_abby_file(input_file, output_file)





